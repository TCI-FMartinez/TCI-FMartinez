
import xlsxwriter

from os import path
from time import sleep
from openpyxl import load_workbook


def mk_excel(ns_dict, archivo="parametros.xlsx"):
    """
    Crea un Excel con la información de la metadata.
    
    - La primera fila contiene los 'keys' (ej.: "N000", "N001", ...).
    - La segunda fila contiene el primer elemento de cada tupla (ej.: "Ruta del parámetro", "Nombre", ...).

    """
    # Obtener las claves en orden (omitimos "properties")
    keys = [k for k in ns_dict if k != "properties"]
    
    # Crear el libro y la hoja usando xlsxwriter
    workbook = xlsxwriter.Workbook(archivo)
    worksheet = workbook.add_worksheet("Hoja1")
    
    # Escribir la primera fila: los keys
    for col, key in enumerate(keys):
        worksheet.write(0, col, key)
    
    # Escribir la segunda fila: el título (primer elemento de la tupla para cada key)
    for col, key in enumerate(keys):
        # Si la tupla está vacía o no tiene elemento en la posición 0, escribimos cadena vacía.
        titulo = ns_dict[key][0] if ns_dict[key] and len(ns_dict[key]) > 0 else ""
        worksheet.write(1, col, titulo)
            # Formato para la primera fila (cabecera)
        formato_fila1 = workbook.add_format({
            'bold': True,
            'italic': False,
            'font_color': 'black'
        })
        worksheet.set_row(0, None, formato_fila1)
        
        # Formato para la segunda fila (descripciones)
        formato_fila2 = workbook.add_format({
            'bold': True,
            'italic': True,
            'font_color': 'black'
        })
        worksheet.set_row(1, None, formato_fila2)

    workbook.close()
    sleep(2)
    print(f"Archivo '{archivo}' creado con la metadata.")

def add_data_lines(archivo, lines, ns_dict):
    """
    Añade nuevas filas al Excel a partir de la fila 3.
    
    Cada fila nueva se compone de las columnas definidas en la metadata (ns_dict).
    Si a la fila le falta algún campo, se completa con cadena vacía.
    
    Parámetros:
      archivo: nombre (o ruta) del archivo Excel.
      lines: lista de diccionarios. Cada diccionario representa una fila de datos.
      ns_dict: diccionario con la metadata, para conocer las columnas esperadas.
    """
    if not path.exists(archivo):
        print(f"El archivo '{archivo}' no existe. Crea el Excel con la metadata primero.")
        return False
    
    # Abrir el libro con openpyxl
    try:
        wb = load_workbook(archivo)
    except Exception as e:
        print("Error al abrir el archivo:", e)
        return False

    if "Hoja1" not in wb.sheetnames:
        print("La hoja 'Hoja1' no existe en el archivo.")
        return False

    ws = wb["Hoja1"]
    
    # Obtener las claves (columnas) en el orden en que fueron escritas en la metadata.
    # Esto se asume que es el mismo orden que en create_metadata_excel.
    keys = [k for k in ns_dict if k != "properties"]
    
    # Recorrer cada nueva línea y preparar la lista de valores en el orden de 'keys'.
    for new_line in lines:
        row_values = [new_line.get(key, "") for key in keys]
        ws.append(row_values)
    
    try:
        wb.save(archivo)
        print(f"Nuevas filas añadidas correctamente al archivo '{archivo}'.")
    except Exception as e:
        print("Error al guardar el archivo:", e)
        return False
    
    return True

# Ejemplo de uso:
if __name__ == "__main__":
    from param_methadata import param_methadata

    ns_dict, _ = param_methadata()

    archivo = "parametros.xlsx"
    
    # Primero creamos el Excel con la metadata:
    mk_excel(ns_dict, archivo)
    
    # Supongamos que las líneas a añadir vienen de otra fuente; por ejemplo:
    lineas = [
        {
            "N000": "ruta/ejemplo_1",
            "N001": "Parámetro 1",
            "N002": "P1",
            "N003": "DIN001",
            "N004": 5.0,
            "N005": 100,
            "N006": 50,
        },
        {
            "N000": "ruta/ejemplo_2",
            "N001": "Parámetro 2",
            # Falta "N002": se completará con cadena vacía.
            "N003": "DIN002",
            "N004": 7.5,
            "N005": 150,
            "N006": 60,
        },
    ]
    
    # Añadir las líneas de datos
    add_data_lines(archivo, lineas, ns_dict)
