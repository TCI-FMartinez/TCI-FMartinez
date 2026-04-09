#!/usr/bin/env python3
import argparse
import json
import os
from collections import Counter, defaultdict
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def safe_mean(values):
    return mean(values) if values else None


def pct(part, total):
    return 0.0 if total == 0 else 100.0 * part / total


def fmt_num(value, digits=2):
    if value is None:
        return "-"
    if isinstance(value, int):
        return str(value)
    return f"{value:.{digits}f}"


def load_rows(path):
    with open(path, "r", encoding="utf-8") as f:
        rows = json.load(f)
    if not isinstance(rows, list):
        raise ValueError("El JSON de entrada debe ser una lista de resultados")

    for row in rows:
        tool_file = row.get("tool_file", "")
        tool_name = os.path.basename(tool_file)
        tool_name = tool_name.replace("_with_polygons.json", "").replace(".json", "")
        row["tool_name"] = tool_name or "unknown_tool"
        row["robot_group"] = "SCARA" if str(row.get("piece_file", "")).startswith("SCARA/") else "ANTHRO"
    return rows


def group_by_piece(rows):
    pieces = defaultdict(list)
    for row in rows:
        key = row.get("piece_reference") or row.get("piece_id") or row.get("piece_file")
        pieces[key].append(row)
    return pieces


def build_stats(rows):
    pieces = group_by_piece(rows)
    tools = sorted({r["tool_name"] for r in rows})

    overview = {
        "total_rows": len(rows),
        "total_pieces": len(pieces),
        "total_tools": len(tools),
        "tools": tools,
        "valid_rows": sum(1 for r in rows if r.get("solution_valid")),
        "invalid_rows": sum(1 for r in rows if not r.get("solution_valid")),
        "status_counts": Counter(r.get("status", "unknown") for r in rows),
        "robot_group_counts": Counter(r.get("robot_group", "UNKNOWN") for r in rows),
    }
    overview["valid_rate"] = pct(overview["valid_rows"], overview["total_rows"])

    tool_stats = {}
    for tool in tools:
        tool_rows = [r for r in rows if r["tool_name"] == tool]
        valid_rows = [r for r in tool_rows if r.get("solution_valid")]
        invalid_rows = [r for r in tool_rows if not r.get("solution_valid")]
        valid_active = [int(r.get("tool_active_count", 0) or 0) for r in valid_rows]
        invalid_active = [int(r.get("tool_active_count", 0) or 0) for r in invalid_rows]
        tool_stats[tool] = {
            "attempts": len(tool_rows),
            "valid": len(valid_rows),
            "invalid": len(invalid_rows),
            "success_rate": pct(len(valid_rows), len(tool_rows)),
            "status_counts": Counter(r.get("status", "unknown") for r in tool_rows),
            "avg_active_valid": safe_mean(valid_active),
            "avg_active_invalid": safe_mean(invalid_active),
            "avg_fxmin_valid": safe_mean([float(r.get("solver_fxmin", 0.0) or 0.0) for r in valid_rows]),
            "avg_fxmin_invalid": safe_mean([float(r.get("solver_fxmin", 0.0) or 0.0) for r in invalid_rows]),
            "robot_group_counts": Counter((r.get("robot_group", "UNKNOWN"), r.get("status", "unknown")) for r in tool_rows),
        }

    piece_stats = []
    for piece_ref, piece_rows in sorted(pieces.items(), key=lambda x: str(x[0])):
        valid_rows = [r for r in piece_rows if r.get("solution_valid")]
        invalid_rows = [r for r in piece_rows if not r.get("solution_valid")]
        best_valid_row = min(valid_rows, key=lambda r: float(r.get("solver_fxmin", 0.0) or 0.0)) if valid_rows else None
        piece_stats.append({
            "piece_reference": piece_ref,
            "piece_id": piece_rows[0].get("piece_id"),
            "piece_file": piece_rows[0].get("piece_file"),
            "robot_group": piece_rows[0].get("robot_group"),
            "valid_count": len(valid_rows),
            "invalid_count": len(invalid_rows),
            "valid_tools": [r["tool_name"] for r in valid_rows],
            "invalid_tools": [r["tool_name"] for r in invalid_rows],
            "has_any_valid": bool(valid_rows),
            "best_valid_tool": best_valid_row["tool_name"] if best_valid_row else None,
            "best_valid_fxmin": best_valid_row.get("solver_fxmin") if best_valid_row else None,
            "status_by_tool": {r["tool_name"]: r.get("status") for r in piece_rows},
        })

    return overview, tool_stats, piece_stats


def derive_recommendations(rows, tool_stats, piece_stats):
    recs = []

    no_global_solution = [p for p in piece_stats if not p["has_any_valid"]]
    if no_global_solution:
        piece_list = ", ".join(p["piece_reference"] for p in no_global_solution)
        recs.append(
            f"Hay piezas sin ninguna solución válida: {piece_list}. Conviene diseñar una herramienta dedicada o una variante compacta para este subconjunto."
        )
    else:
        recs.append(
            "No hay piezas sin solución global. Todos los casos tienen al menos una herramienta válida; los fallos son específicos de ciertas herramientas."
        )

    cannot_lift = [r for r in rows if r.get("status") == "infeasible_cannot_lift"]
    no_fit = [r for r in rows if int(r.get("solver_error_flag", 999)) == -5 or r.get("status") == "solver_error"]

    if cannot_lift:
        saturated = [
            r for r in cannot_lift
            if int(r.get("tool_active_count", 0) or 0) >= int(r.get("tool_elements_total", 0) or 0)
            and int(r.get("tool_elements_total", 0) or 0) > 0
        ]
        if saturated:
            recs.append(
                "Hay varios fallos por 'cannot_lift' con la herramienta completamente activada. Esto apunta a falta de capacidad de elevación o mala distribución de apoyo, no a falta de cobertura geométrica."
            )
            recs.append(
                "Para esas herramientas conviene aumentar la capacidad útil por punto, redistribuir los apoyos hacia zonas con más brazo resistente y añadir más puntos de agarre efectivos en la periferia o en zonas de mayor momento."
            )
        else:
            recs.append(
                "Los fallos por 'cannot_lift' aparecen sin saturar todos los actuadores. Tiene sentido revisar tanto la lógica de selección como la distribución real de puntos activos."
            )

    if no_fit:
        recs.append(
            "Los casos con flag -5 ('no actuator fits') indican un problema geométrico: paso demasiado grande, huella poco adaptable o interferencia con la pieza."
        )
        recs.append(
            "Para estos casos conviene crear una variante compacta: menor paso entre actuadores, filas desplazadas, módulos más pequeños o una subherramienta específica para piezas estrechas o con zonas útiles muy localizadas."
        )

    for tool_name, stats in sorted(tool_stats.items(), key=lambda x: x[1]["success_rate"]):
        status_counts = stats["status_counts"]
        if stats["success_rate"] < 70.0:
            if status_counts.get("infeasible_cannot_lift", 0) >= status_counts.get("solver_error", 0):
                recs.append(
                    f"{tool_name}: su problema dominante es la elevación. Prioridad de rediseño: aumentar capacidad y rigidez del patrón de agarre antes que añadir más volumen global."
                )
            else:
                recs.append(
                    f"{tool_name}: su problema dominante es geométrico. Prioridad de rediseño: compactar la matriz y mejorar accesibilidad a zonas pequeñas."
                )

    replacement_patterns = Counter()
    for piece in piece_stats:
        if piece["has_any_valid"] and piece["invalid_tools"]:
            replacement_patterns[(tuple(sorted(piece["invalid_tools"])), tuple(sorted(piece["valid_tools"])))] += 1
    if replacement_patterns:
        (invalid_tools, valid_tools), count = replacement_patterns.most_common(1)[0]
        recs.append(
            f"Patrón dominante detectado en {count} piezas: fallan {', '.join(invalid_tools)} y resuelve {', '.join(valid_tools)}. Esto sugiere usar esa familia válida como referencia de rediseño."
        )

    return recs


def build_markdown(overview, tool_stats, piece_stats, recommendations):
    lines = []
    lines.append("# Informe de resultados de herramientas")
    lines.append("")
    lines.append("## Resumen general")
    lines.append("")
    lines.append(f"- Combinaciones evaluadas: {overview['total_rows']}")
    lines.append(f"- Piezas únicas: {overview['total_pieces']}")
    lines.append(f"- Herramientas evaluadas: {', '.join(overview['tools'])}")
    lines.append(f"- Soluciones válidas: {overview['valid_rows']} ({overview['valid_rate']:.2f}%)")
    lines.append(f"- Soluciones no válidas: {overview['invalid_rows']} ({pct(overview['invalid_rows'], overview['total_rows']):.2f}%)")
    lines.append("")
    lines.append("### Estados")
    lines.append("")
    for status, count in sorted(overview["status_counts"].items()):
        lines.append(f"- {status}: {count}")
    lines.append("")
    lines.append("## Rendimiento por herramienta")
    lines.append("")
    lines.append("| Herramienta | Intentos | Válidas | No válidas | Éxito % | Activos medios válidos | Activos medios no válidos |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|")
    for tool, stats in sorted(tool_stats.items(), key=lambda x: (-x[1]["success_rate"], x[0])):
        lines.append(
            f"| {tool} | {stats['attempts']} | {stats['valid']} | {stats['invalid']} | {stats['success_rate']:.2f} | {fmt_num(stats['avg_active_valid'])} | {fmt_num(stats['avg_active_invalid'])} |"
        )
    lines.append("")

    lines.append("## Estado por pieza")
    lines.append("")
    lines.append("| Pieza | Grupo | Herramientas válidas | Herramientas no válidas | Mejor herramienta válida |")
    lines.append("|---|---|---|---|---|")
    for piece in piece_stats:
        lines.append(
            f"| {piece['piece_reference']} | {piece['robot_group']} | {', '.join(piece['valid_tools']) if piece['valid_tools'] else '-'} | {', '.join(piece['invalid_tools']) if piece['invalid_tools'] else '-'} | {piece['best_valid_tool'] or '-'} |"
        )
    lines.append("")

    lines.append("## Recomendaciones")
    lines.append("")
    for rec in recommendations:
        lines.append(f"- {rec}")
    lines.append("")

    only_one_valid = [p for p in piece_stats if p['valid_count'] == 1]
    if only_one_valid:
        lines.append("## Piezas con una única herramienta válida")
        lines.append("")
        for piece in only_one_valid:
            lines.append(
                f"- {piece['piece_reference']}: válida {piece['best_valid_tool']}; fallan {', '.join(piece['invalid_tools'])}"
            )
        lines.append("")

    return "\n".join(lines)


def write_json_summary(path, overview, tool_stats, piece_stats, recommendations):
    payload = {
        "overview": {
            **overview,
            "status_counts": dict(overview["status_counts"]),
            "robot_group_counts": dict(overview["robot_group_counts"]),
        },
        "tool_stats": {
            tool: {
                **stats,
                "status_counts": dict(stats["status_counts"]),
                "robot_group_counts": {f"{k[0]}::{k[1]}": v for k, v in stats["robot_group_counts"].items()},
            }
            for tool, stats in tool_stats.items()
        },
        "piece_stats": piece_stats,
        "recommendations": recommendations,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _apply_header_style(ws, row_idx=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, min_width=12, max_width=48):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))
    for column_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(column_idx)].width = max(min_width, min(max_width, width + 2))


def _write_sheet_rows(ws, headers, rows):
    ws.append(headers)
    _apply_header_style(ws, 1)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    _autosize_columns(ws)


def write_report_workbook(path, overview, tool_stats, piece_stats, recommendations, rows):
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Resumen"

    overview_rows = [
        ["total_rows", overview["total_rows"]],
        ["total_pieces", overview["total_pieces"]],
        ["total_tools", overview["total_tools"]],
        ["valid_rows", overview["valid_rows"]],
        ["invalid_rows", overview["invalid_rows"]],
        ["valid_rate_pct", overview["valid_rate"] / 100.0],
    ]
    overview_rows.extend([[f"status::{status}", count] for status, count in sorted(overview["status_counts"].items())])
    overview_rows.extend([[f"robot::{robot}", count] for robot, count in sorted(overview["robot_group_counts"].items())])
    _write_sheet_rows(ws_overview, ["metric", "value"], overview_rows)
    ws_overview["B7"].number_format = "0.0%"
    for cell in ws_overview["B"]:
        if cell.row == 1:
            continue
        if isinstance(cell.value, (int, float)) and ws_overview.cell(cell.row, 1).value == "valid_rate_pct":
            cell.number_format = "0.0%"

    ws_tools = wb.create_sheet("Herramientas")
    tool_rows = []
    for tool, stats in sorted(tool_stats.items(), key=lambda x: (-x[1]["success_rate"], x[0])):
        tool_rows.append([
            tool,
            stats["attempts"],
            stats["valid"],
            stats["invalid"],
            stats["success_rate"] / 100.0,
            stats["avg_active_valid"],
            stats["avg_active_invalid"],
            stats["avg_fxmin_valid"],
            stats["avg_fxmin_invalid"],
            "; ".join(f"{k}={v}" for k, v in sorted(stats["status_counts"].items())),
            "; ".join(f"{robot}/{status}={count}" for (robot, status), count in sorted(stats["robot_group_counts"].items())),
        ])
    _write_sheet_rows(
        ws_tools,
        ["tool_name", "attempts", "valid", "invalid", "success_rate", "avg_active_valid", "avg_active_invalid", "avg_fxmin_valid", "avg_fxmin_invalid", "status_counts", "robot_group_counts"],
        tool_rows,
    )
    for row in range(2, ws_tools.max_row + 1):
        ws_tools.cell(row, 5).number_format = "0.0%"

    ws_pieces = wb.create_sheet("Piezas")
    piece_rows = []
    for piece in piece_stats:
        piece_rows.append([
            piece["piece_reference"],
            piece["piece_id"],
            piece["robot_group"],
            piece["piece_file"],
            piece["has_any_valid"],
            piece["valid_count"],
            piece["invalid_count"],
            ";".join(piece["valid_tools"]),
            ";".join(piece["invalid_tools"]),
            piece["best_valid_tool"] or "",
            piece["best_valid_fxmin"],
            "; ".join(f"{tool}={status}" for tool, status in sorted(piece["status_by_tool"].items())),
        ])
    _write_sheet_rows(
        ws_pieces,
        ["piece_reference", "piece_id", "robot_group", "piece_file", "has_any_valid", "valid_count", "invalid_count", "valid_tools", "invalid_tools", "best_valid_tool", "best_valid_fxmin", "status_by_tool"],
        piece_rows,
    )

    ws_recs = wb.create_sheet("Recomendaciones")
    rec_rows = [[idx, rec] for idx, rec in enumerate(recommendations, start=1)]
    _write_sheet_rows(ws_recs, ["orden", "recomendacion"], rec_rows)
    ws_recs.column_dimensions["B"].width = 120
    for row in range(2, ws_recs.max_row + 1):
        ws_recs.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    ws_raw = wb.create_sheet("Raw")
    raw_headers = [
        "robot_group", "piece_reference", "piece_id", "piece_file", "tool_name", "status", "solution_valid",
        "tool_active_count", "tool_elements_total", "solver_fxmin", "solver_error_flag", "returncode_signed",
        "center_distance_approx", "solution_json", "combo_dir"
    ]
    raw_rows = []
    for row in rows:
        raw_rows.append([
            row.get("robot_group"),
            row.get("piece_reference"),
            row.get("piece_id"),
            row.get("piece_file"),
            row.get("tool_name"),
            row.get("status"),
            row.get("solution_valid"),
            row.get("tool_active_count"),
            row.get("tool_elements_total"),
            row.get("solver_fxmin"),
            row.get("solver_error_flag"),
            row.get("returncode_signed"),
            row.get("center_distance_approx"),
            row.get("solution_json"),
            row.get("combo_dir"),
        ])
    _write_sheet_rows(ws_raw, raw_headers, raw_rows)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def generate_tool_report_files(input_json="summary.json", output_dir="report_out", base_name="tool_report"):
    e01 = e02 = None
    mess1 = mess2 = None
    rows = load_rows(input_json)
    overview, tool_stats, piece_stats = build_stats(rows)
    recommendations = derive_recommendations(rows, tool_stats, piece_stats)

    os.makedirs(output_dir, exist_ok=True)
    md_path = os.path.join(output_dir, f"{base_name}.md")
    xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
    json_path = os.path.join(output_dir, f"{base_name}_summary.json")

    markdown = build_markdown(overview, tool_stats, piece_stats, recommendations)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(markdown)
    try:
        write_report_workbook(xlsx_path, overview, tool_stats, piece_stats, recommendations, rows)
    except Exception as e01:
        mess1= (f"Error al generar el informe Excel: {e01}")
    try:
        write_json_summary(json_path, overview, tool_stats, piece_stats, recommendations)
    except Exception as e02:
        mess2 = (f"Error al generar el resumen JSON: {e02}")

    return mess1, mess2


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Genera un informe a partir de summary.json")
    parser.add_argument("input_json", nargs="?", default="summary.json", help="Ruta al summary.json")
    parser.add_argument("--output-dir", default="report_out", help="Carpeta de salida")
    parser.add_argument("--base-name", default="tool_report", help="Prefijo de los archivos de salida")
    args = parser.parse_args()
    generate_tool_report_files(args.input_json, args.output_dir, args.base_name)